from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Depends
from fastapi.responses import StreamingResponse, Response
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import urllib.request
from jose import JWTError, jwt
from passlib.context import CryptContext
import sys

ROOT_DIR = Path(__file__).parent
sys.path.insert(0, str(ROOT_DIR))
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Add CORS middleware (MUST be added before routes)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# JWT Settings
SECRET_KEY = os.environ.get('JWT_SECRET', 'thryve-coworking-secret-key-2026')  # Set JWT_SECRET in .env for production
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_HOURS = 24

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# Security
security = HTTPBearer()

# GST Rate
GST_RATE = 18  # 18% GST (9% CGST + 9% SGST or 18% IGST)

# Thryve Coworking Company Details
COMPANY_DETAILS = {
    "name": "Thryve Coworking",
    "address": "18/1, Plot no. 3, Azad Colony, Mathura Road, Sector 15 A, Faridabad, Haryana 121007",
    "gstin": "06AAYFT8213A1Z2",
    "state": "Haryana",
    "email": "billing@thryvecoworking.com",
    "phone": "+91 80 1234 5678"
}

# User Roles
ROLES = {
    "admin": ["all"],  # Full access
    "staff": ["create_invoice", "view_invoice", "view_client", "create_client", "download_pdf", "bulk_invoice"],
    "viewer": ["view_invoice", "view_client", "download_pdf"]
}

# Auth Models
class UserCreate(BaseModel):
    name: str
    email: str
    password: str
    role: str = "staff"  # admin, staff, viewer

class UserLogin(BaseModel):
    email: str
    password: str

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    email: str
    password_hash: str
    role: str = "staff"
    is_active: bool = True
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    created_by: Optional[str] = None

class UserResponse(BaseModel):
    id: str
    name: str
    email: str
    role: str
    is_active: bool
    created_at: str

# Auth Helper Functions
def verify_password(plain_password: str, hashed_password: str) -> bool:
    try:
        result = pwd_context.verify(plain_password, hashed_password)
        print(f"[DEBUG] verify_password: plain='{plain_password}', hash='{hashed_password[:20]}...', result={result}")
        return result
    except Exception as e:
        print(f"[DEBUG] verify_password ERROR: {e}")
        return False

def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(hours=ACCESS_TOKEN_EXPIRE_HOURS)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    token = credentials.credentials
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if user is None:
        raise HTTPException(status_code=401, detail="User not found")
    if not user.get("is_active", True):
        raise HTTPException(status_code=401, detail="User is inactive")
    return user

def check_permission(user: dict, permission: str):
    role = user.get("role", "viewer")
    permissions = ROLES.get(role, [])
    if "all" in permissions or permission in permissions:
        return True
    raise HTTPException(status_code=403, detail="Permission denied")

# Define Models
class ClientBase(BaseModel):
    company_name: str
    address: str
    gstin: str

class ClientCreate(ClientBase):
    pass

class Client(ClientBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class LineItem(BaseModel):
    description: str
    service_type: str  # monthly_rental, security_deposit, setup_charges, day_pass, meeting_room
    quantity: float = 1
    rate: float
    is_taxable: bool = True  # GST applicable or not
    hsn_sac: Optional[str] = None  # HSN/SAC code
    unit: str = "Units"  # Units, Month, Day, Hour
    # Prorate fields
    is_prorated: bool = False
    prorate_days: Optional[int] = None  # Number of days to charge
    prorate_total_days: Optional[int] = None  # Total days in billing period (usually 30)

class InvoiceLineItem(LineItem):
    amount: float = 0
    cgst: float = 0
    sgst: float = 0
    total: float = 0

class InvoiceCreate(BaseModel):
    client_id: str
    invoice_date: str
    due_date: Optional[str] = None
    line_items: List[LineItem]
    notes: Optional[str] = ""

class InvoiceStatusUpdate(BaseModel):
    status: str  # pending, paid, overdue
    payment_date: Optional[str] = None

class Invoice(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    invoice_number: str
    invoice_date: str
    due_date: Optional[str] = None
    client: dict
    line_items: List[dict]
    subtotal: float
    total_cgst: float
    total_sgst: float
    total_tax: float
    round_off_adjustment: float = 0  # GST-compliant rounding adjustment
    grand_total: float
    notes: str = ""
    status: str = "pending"  # pending, paid, overdue
    payment_date: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    company: dict = Field(default_factory=lambda: COMPANY_DETAILS)

# Helper function to generate invoice number
# Format: YYYY-YYYY/MM/SEQ/CLIENT (e.g., 2026-2027/04/001/Himanshu)
async def generate_invoice_number(company_name: str = "", invoice_date: str = None):
    """
    Generate invoice number in format: YYYY-YYYY/MM/NNN/CompanyName
    - YYYY-YYYY: Financial year (April to March), e.g., 2026-2027
    - MM: Two-digit month (01-12)
    - NNN: Sequential number, resets each financial year
    - CompanyName: Client company name for easy identification
    
    Financial year is derived from the invoice date:
    - April to December: Current Year - Next Year (e.g., May 2026 → 2026-2027)
    - January to March: Previous Year - Current Year (e.g., March 2026 → 2025-2026)
    """
    # Use invoice date if provided, otherwise use current date
    if invoice_date:
        try:
            date_obj = datetime.strptime(invoice_date.split('T')[0], "%Y-%m-%d")
        except:
            date_obj = datetime.now()
    else:
        date_obj = datetime.now()
    
    # Determine financial year (April to March) from invoice date
    if date_obj.month >= 4:  # April onwards
        fy_start_year = date_obj.year
        fy_end_year = date_obj.year + 1
    else:  # Jan-Mar
        fy_start_year = date_obj.year - 1
        fy_end_year = date_obj.year
    
    # Format: 2026-2027 for FY 2026-2027
    fy_prefix = f"{fy_start_year}-{fy_end_year}"
    
    # Two-digit month from invoice date
    month_num = f"{date_obj.month:02d}"
    
    # Query for invoices in this financial year using the new format
    fy_pattern = f"^{fy_prefix}/"
    latest = await db.invoices.find_one(
        {"invoice_number": {"$regex": fy_pattern}},
        {"invoice_number": 1},
        sort=[("created_at", -1), ("invoice_number", -1)]
    )
    
    if latest:
        # Extract the sequence number (third part after splitting by /)
        try:
            parts = latest["invoice_number"].split("/")
            if len(parts) >= 3:
                last_num = int(parts[2])
                next_num = last_num + 1
            else:
                next_num = 1
        except:
            next_num = 1
    else:
        next_num = 1
    
    # Clean company name for invoice number (remove special chars, limit length)
    clean_company = company_name.strip()
    if len(clean_company) > 30:
        clean_company = clean_company[:30].strip()
    
    # Format: 2026-2027/04/001/Company Name
    return f"{fy_prefix}/{month_num}/{next_num:03d}/{clean_company}"

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/register")
async def register_user(user_data: UserCreate, current_user: dict = Depends(get_current_user)):
    # Only admin can create users
    check_permission(current_user, "all")
    
    # Check if email already exists
    existing = await db.users.find_one({"email": user_data.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Validate role
    if user_data.role not in ROLES:
        raise HTTPException(status_code=400, detail="Invalid role. Must be: admin, staff, or viewer")
    
    # Create user
    user = User(
        name=user_data.name,
        email=user_data.email.lower(),
        password_hash=get_password_hash(user_data.password),
        role=user_data.role,
        created_by=current_user.get("id")
    )
    
    doc = user.model_dump()
    await db.users.insert_one(doc)
    
    return {"message": "User created successfully", "user_id": user.id}

@api_router.post("/auth/login")
async def login(user_data: UserLogin):
    print(f"[LOGIN] Attempting login for: {user_data.email}")
    user = await db.users.find_one({"email": user_data.email.lower()}, {"_id": 0})
    if not user:
        print(f"[LOGIN] User not found: {user_data.email}")
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    print(f"[LOGIN] User found: {user.get('email')}")
    password_valid = verify_password(user_data.password, user["password_hash"])
    print(f"[LOGIN] Password valid: {password_valid}")
    
    if not password_valid:
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    if not user.get("is_active", True):
        raise HTTPException(status_code=401, detail="Account is inactive")
    
    # Create access token
    access_token = create_access_token(data={"sub": user["id"], "role": user["role"]})
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user": {
            "id": user["id"],
            "name": user["name"],
            "email": user["email"],
            "role": user["role"]
        }
    }

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    return {
        "id": current_user["id"],
        "name": current_user["name"],
        "email": current_user["email"],
        "role": current_user["role"]
    }

@api_router.get("/auth/users", response_model=List[UserResponse])
async def get_users(current_user: dict = Depends(get_current_user)):
    check_permission(current_user, "all")
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(100)
    return users

@api_router.delete("/auth/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_current_user)):
    check_permission(current_user, "all")
    
    if user_id == current_user["id"]:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "User deleted successfully"}

@api_router.patch("/auth/users/{user_id}/toggle")
async def toggle_user_status(user_id: str, current_user: dict = Depends(get_current_user)):
    check_permission(current_user, "all")
    
    if user_id == current_user["id"]:
        raise HTTPException(status_code=400, detail="Cannot deactivate yourself")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    new_status = not user.get("is_active", True)
    await db.users.update_one({"id": user_id}, {"$set": {"is_active": new_status}})
    
    return {"message": f"User {'activated' if new_status else 'deactivated'}", "is_active": new_status}

class PasswordChange(BaseModel):
    current_password: str
    new_password: str

@api_router.post("/auth/change-password")
async def change_password(data: PasswordChange, current_user: dict = Depends(get_current_user)):
    # Verify current password
    user = await db.users.find_one({"id": current_user["id"]}, {"_id": 0})
    if not verify_password(data.current_password, user["password_hash"]):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    
    # Update password
    new_hash = get_password_hash(data.new_password)
    await db.users.update_one({"id": current_user["id"]}, {"$set": {"password_hash": new_hash}})
    
    return {"message": "Password changed successfully"}

# Create default admin user on startup
@app.on_event("startup")
async def create_default_admin():
    admin = await db.users.find_one({"role": "admin"})
    if not admin:
        default_admin = User(
            name="Admin",
            email="admin@thryve.in",
            password_hash=get_password_hash("password"),
            role="admin"
        )
        await db.users.insert_one(default_admin.model_dump())
        logger.info("Default admin user created: admin@thryve.in / password")
    
    # Seed default management data
    from routes.management import seed_default_data
    await seed_default_data(db)
    
    # Seed default public holidays
    await seed_default_holidays(db)

# Calculate line item with GST
def calculate_line_item(item: LineItem) -> dict:
    # Handle prorate calculation
    if item.is_prorated and item.prorate_days and item.prorate_total_days:
        prorate_rate = (item.rate / item.prorate_total_days) * item.prorate_days
        amount = item.quantity * prorate_rate
    else:
        amount = item.quantity * item.rate
    
    if item.is_taxable:
        cgst = round(amount * (GST_RATE / 2) / 100, 2)
        sgst = round(amount * (GST_RATE / 2) / 100, 2)
    else:
        cgst = 0
        sgst = 0
    total = amount + cgst + sgst
    return {
        "description": item.description,
        "service_type": item.service_type,
        "quantity": item.quantity,
        "rate": item.rate,
        "is_taxable": item.is_taxable,
        "hsn_sac": item.hsn_sac or "",
        "unit": item.unit,
        "is_prorated": item.is_prorated,
        "prorate_days": item.prorate_days,
        "prorate_total_days": item.prorate_total_days,
        "amount": round(amount, 2),
        "cgst": cgst,
        "sgst": sgst,
        "total": round(total, 2)
    }

# API Routes

@api_router.get("/")
async def root():
    return {"message": "Thryve Invoice Generator API"}

@api_router.get("/company")
async def get_company_details():
    return COMPANY_DETAILS

# Client CRUD
@api_router.post("/clients", response_model=Client)
async def create_client(client_data: ClientCreate):
    client_obj = Client(**client_data.model_dump())
    doc = client_obj.model_dump()
    await db.clients.insert_one(doc)
    return client_obj

@api_router.get("/clients", response_model=List[Client])
async def get_clients():
    clients = await db.clients.find({}, {"_id": 0}).to_list(1000)
    return clients

@api_router.get("/clients/{client_id}", response_model=Client)
async def get_client(client_id: str):
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    return client

@api_router.put("/clients/{client_id}", response_model=Client)
async def update_client(client_id: str, client_data: ClientCreate):
    existing = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Client not found")
    
    update_data = client_data.model_dump()
    await db.clients.update_one({"id": client_id}, {"$set": update_data})
    
    updated = await db.clients.find_one({"id": client_id}, {"_id": 0})
    return updated

@api_router.delete("/clients/{client_id}")
async def delete_client(client_id: str):
    result = await db.clients.delete_one({"id": client_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Client not found")
    return {"message": "Client deleted successfully"}

# Invoice CRUD
@api_router.post("/invoices", response_model=Invoice)
async def create_invoice(invoice_data: InvoiceCreate):
    # Get client details - check both clients and companies collections
    client = await db.clients.find_one({"id": invoice_data.client_id}, {"_id": 0})
    if not client:
        # Try companies collection
        company = await db.companies.find_one({"id": invoice_data.client_id}, {"_id": 0})
        if company:
            # Map company fields to client format for invoice
            client = {
                "id": company["id"],
                "name": company.get("signatory_name", ""),
                "company_name": company["company_name"],
                "address": company.get("company_address", ""),
                "gstin": company.get("company_gstin", ""),
                "email": company.get("signatory_email", ""),
                "phone": company.get("signatory_phone", "")
            }
        else:
            raise HTTPException(status_code=404, detail="Client not found")
    
    # Generate invoice number with company name and invoice date
    company_name = client.get("company_name") or client.get("name", "")
    invoice_number = await generate_invoice_number(company_name, invoice_data.invoice_date)
    
    # Calculate line items and collect booking IDs for meeting room charges
    calculated_items = []
    all_booking_ids = []
    
    for item in invoice_data.line_items:
        calculated_item = calculate_line_item(item)
        calculated_items.append(calculated_item)
        
        # Collect booking IDs from bundled meeting room charges
        if hasattr(item, 'booking_ids') and item.booking_ids:
            all_booking_ids.extend(item.booking_ids)
        elif isinstance(item, dict) and item.get('booking_ids'):
            all_booking_ids.extend(item.get('booking_ids', []))
    
    # Calculate totals with full precision
    subtotal = sum(item["amount"] for item in calculated_items)
    total_cgst = sum(item["cgst"] for item in calculated_items)
    total_sgst = sum(item["sgst"] for item in calculated_items)
    total_tax = total_cgst + total_sgst
    calculated_total = round(subtotal + total_tax, 2)
    
    # GST-compliant rounding: Round final total to nearest whole rupee
    rounded_total = round(calculated_total)
    round_off_adjustment = round(rounded_total - calculated_total, 2)
    
    # Create invoice object
    invoice_obj = Invoice(
        invoice_number=invoice_number,
        invoice_date=invoice_data.invoice_date,
        due_date=invoice_data.due_date,
        client=client,
        line_items=calculated_items,
        subtotal=round(subtotal, 2),
        total_cgst=round(total_cgst, 2),
        total_sgst=round(total_sgst, 2),
        total_tax=round(total_tax, 2),
        round_off_adjustment=round_off_adjustment,
        grand_total=rounded_total,
        notes=invoice_data.notes or "",
        status="pending"
    )
    
    doc = invoice_obj.model_dump()
    # Remove None values to avoid storing them
    doc = {k: v for k, v in doc.items() if v is not None}
    
    # Store booking IDs if present (for tracking)
    if all_booking_ids:
        doc["meeting_room_booking_ids"] = all_booking_ids
    
    await db.invoices.insert_one(doc)
    
    # Mark associated bookings as invoiced
    if all_booking_ids:
        await db.bookings.update_many(
            {"id": {"$in": all_booking_ids}},
            {"$set": {
                "payment_status": "invoiced",
                "invoice_id": invoice_obj.id,
                "invoiced_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    
    return invoice_obj

@api_router.get("/invoices", response_model=List[Invoice])
async def get_invoices():
    invoices = await db.invoices.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return invoices

@api_router.get("/invoices/{invoice_id}", response_model=Invoice)
async def get_invoice(invoice_id: str):
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return invoice

@api_router.delete("/invoices/{invoice_id}")
async def delete_invoice(invoice_id: str):
    result = await db.invoices.delete_one({"id": invoice_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return {"message": "Invoice deleted successfully"}


class InvoiceUpdate(BaseModel):
    """Update invoice - everything except invoice_number and invoice_date"""
    client_id: Optional[str] = None
    due_date: Optional[str] = None
    line_items: Optional[List[dict]] = None
    notes: Optional[str] = None
    status: Optional[str] = None


@api_router.put("/invoices/{invoice_id}", response_model=Invoice)
async def update_invoice(invoice_id: str, update_data: InvoiceUpdate):
    # Get existing invoice
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    update_fields = {}
    
    # Update client if provided
    if update_data.client_id:
        client = await db.clients.find_one({"id": update_data.client_id}, {"_id": 0})
        if not client:
            # Try companies collection
            client = await db.companies.find_one({"id": update_data.client_id}, {"_id": 0})
            if client:
                # Map company fields to client format
                client = {
                    "id": client["id"],
                    "name": client.get("signatory_name", ""),
                    "company_name": client["company_name"],
                    "address": client.get("company_address", ""),
                    "gstin": client.get("company_gstin", ""),
                    "email": client.get("signatory_email", ""),
                    "phone": client.get("signatory_phone", "")
                }
        if not client:
            raise HTTPException(status_code=404, detail="Client not found")
        update_fields["client"] = client
    
    # Update due date
    if update_data.due_date:
        update_fields["due_date"] = update_data.due_date
    
    # Update notes
    if update_data.notes is not None:
        update_fields["notes"] = update_data.notes
    
    # Update status
    if update_data.status:
        update_fields["status"] = update_data.status
    
    # Update line items and recalculate totals
    if update_data.line_items is not None:
        # Convert dict items to LineItem objects for calculation
        line_item_objects = []
        for item in update_data.line_items:
            if isinstance(item, dict):
                line_item_objects.append(LineItem(**item))
            else:
                line_item_objects.append(item)
        
        calculated_items = [calculate_line_item(li) for li in line_item_objects]
        
        subtotal = sum(item["amount"] for item in calculated_items)
        total_cgst = sum(item["cgst"] for item in calculated_items)
        total_sgst = sum(item["sgst"] for item in calculated_items)
        total_tax = total_cgst + total_sgst
        grand_total = subtotal + total_tax
        
        update_fields["line_items"] = calculated_items
        update_fields["subtotal"] = round(subtotal, 2)
        update_fields["total_cgst"] = round(total_cgst, 2)
        update_fields["total_sgst"] = round(total_sgst, 2)
        update_fields["total_tax"] = round(total_tax, 2)
        update_fields["grand_total"] = round(grand_total, 2)
    
    # Add updated timestamp
    update_fields["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    # Perform update
    await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": update_fields}
    )
    
    # Return updated invoice
    updated_invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    return updated_invoice

# Number to words conversion for invoice
def number_to_words(num):
    if num == 0:
        return "Zero"
    
    ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine", "Ten",
        "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
    
    def convert_less_than_thousand(n):
        if n == 0:
            return ""
        if n < 20:
            return ones[n]
        if n < 100:
            return tens[n // 10] + (" " + ones[n % 10] if n % 10 != 0 else "")
        return ones[n // 100] + " Hundred" + (" " + convert_less_than_thousand(n % 100) if n % 100 != 0 else "")
    
    def convert(n):
        if n == 0:
            return ""
        result = ""
        if n >= 10000000:
            result += convert_less_than_thousand(n // 10000000) + " Crore "
            n %= 10000000
        if n >= 100000:
            result += convert_less_than_thousand(n // 100000) + " Lakh "
            n %= 100000
        if n >= 1000:
            result += convert_less_than_thousand(n // 1000) + " Thousand "
            n %= 1000
        if n > 0:
            result += convert_less_than_thousand(n)
        return result.strip()
    
    int_part = int(num)
    return "Rupees " + convert(int_part) + " Only"

# Generate PDF for invoice - WYSIWYG using Playwright
@api_router.get("/invoices/{invoice_id}/pdf")
async def generate_invoice_pdf(invoice_id: str):
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    try:
        # Use fpdf2-based PDF generator (pure Python, no system deps, works in production)
        from utils.pdf_generator import generate_pdf_from_html
        
        pdf_content = generate_pdf_from_html(invoice)
        
        # Create filename
        invoice_num = invoice.get('invoice_number', 'invoice').replace('/', '-')
        client_name = invoice.get('client', {}).get('company_name', '').replace(' ', '_').replace('/', '-').replace('\\', '-')
        client_name = ''.join(c for c in client_name if c.isalnum() or c in ['_', '-'])
        filename = f"{invoice_num}_{client_name}.pdf"
        
        return Response(
            content=pdf_content,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"inline; filename={filename}",
                "Content-Length": str(len(pdf_content))
            }
        )
    except Exception as e:
        print(f"PDF generation error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to generate PDF: {str(e)}")


@api_router.patch("/invoices/{invoice_id}/status", response_model=Invoice)
async def update_invoice_status(invoice_id: str, status_update: InvoiceStatusUpdate):
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    update_data = {"status": status_update.status}
    if status_update.status == "paid" and status_update.payment_date:
        update_data["payment_date"] = status_update.payment_date
    elif status_update.status == "paid":
        update_data["payment_date"] = datetime.now(timezone.utc).isoformat().split('T')[0]
    
    await db.invoices.update_one({"id": invoice_id}, {"$set": update_data})
    updated = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    return updated

# Dashboard stats
@api_router.get("/stats")
async def get_stats():
    total_invoices = await db.invoices.count_documents({})
    total_clients = await db.clients.count_documents({})
    
    # Calculate totals by status
    invoices = await db.invoices.find({}, {"_id": 0, "grand_total": 1, "status": 1, "due_date": 1}).to_list(1000)
    total_revenue = sum(inv.get("grand_total", 0) for inv in invoices)
    
    # Count by status
    paid_count = sum(1 for inv in invoices if inv.get("status") == "paid")
    pending_count = sum(1 for inv in invoices if inv.get("status") == "pending")
    overdue_count = sum(1 for inv in invoices if inv.get("status") == "overdue")
    
    # Calculate pending/overdue amounts
    pending_amount = sum(inv.get("grand_total", 0) for inv in invoices if inv.get("status") in ["pending", "overdue"])
    paid_amount = sum(inv.get("grand_total", 0) for inv in invoices if inv.get("status") == "paid")
    
    return {
        "total_invoices": total_invoices,
        "total_clients": total_clients,
        "total_revenue": round(total_revenue, 2),
        "paid_count": paid_count,
        "pending_count": pending_count,
        "overdue_count": overdue_count,
        "pending_amount": round(pending_amount, 2),
        "paid_amount": round(paid_amount, 2)
    }

# Check and update overdue invoices
@api_router.post("/invoices/check-overdue")
async def check_overdue_invoices():
    today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    # Find pending invoices with due_date before today
    result = await db.invoices.update_many(
        {"status": "pending", "due_date": {"$lt": today}},
        {"$set": {"status": "overdue"}}
    )
    return {"updated_count": result.modified_count}

# Download Excel Template
@api_router.get("/excel/template")
async def download_excel_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Data"
    
    # Define headers - One row per client with all service types as columns
    headers = [
        "Client Company Name", "Client Address", "Client GSTIN",
        "Invoice Date (YYYY-MM-DD)", "Due Date (YYYY-MM-DD)",
        "Monthly Plan Fee", "Prorate Days", "Prorate Total Days",
        "Day Pass Rate", "Day Pass Qty",
        "Security Deposit",
        "Setup Charges",
        "Meeting Room Rate", "Meeting Room Hours",
        "Additional Charges", "Additional Charges Description",
        "Notes"
    ]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E375B", end_color="2E375B", fill_type="solid")
    header_fill_orange = PatternFill(start_color="FFA14A", end_color="FFA14A", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Column categories for coloring
    client_cols = [1, 2, 3]  # Client info
    date_cols = [4, 5]  # Dates
    service_cols = [6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]  # Services
    notes_col = [17]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        if col in service_cols:
            cell.fill = header_fill_orange
            cell.font = Font(bold=True, color="2E375B")
        else:
            cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = 18
    
    # Wider columns for address and description
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['P'].width = 25
    ws.column_dimensions['Q'].width = 30
    
    # Add sample data rows
    sample_data = [
        # Client with monthly plan + security deposit
        ["ABC Tech Pvt Ltd", "123 Tech Park, Sector 15, Faridabad, Haryana 121007", "06AABCT1234F1Z5",
         "2026-01-01", "2026-01-15",
         "15000", "", "",  # Monthly plan (full month)
         "", "",  # No day pass
         "30000",  # Security deposit
         "",  # No setup
         "", "",  # No meeting room
         "", "",  # No additional
         "New member - January billing"],
        # Client with prorated plan + setup
        ["XYZ Solutions", "456 Business Hub, Sector 21, Faridabad, Haryana 121002", "06BBXYZ5678G1Z8",
         "2026-01-15", "2026-01-31",
         "12000", "15", "30",  # Prorated (15 of 30 days)
         "", "",  # No day pass
         "",  # No security
         "5000",  # Setup charges
         "", "",  # No meeting room
         "", "",  # No additional
         "Mid-month joiner"],
        # Client with day pass + meeting room
        ["Startup Inc", "789 Innovation Center, Faridabad, Haryana 121003", "06CCSTU9012H1Z2",
         "2026-01-20", "2026-01-25",
         "", "", "",  # No monthly
         "500", "3",  # 3 day passes
         "",  # No security
         "",  # No setup
         "1000", "4",  # 4 hours meeting room
         "2000", "Printing & Courier",  # Additional charges
         "Guest usage"],
    ]
    
    for row_num, row_data in enumerate(sample_data, 2):
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
    
    # Add instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    instructions = [
        ["THRYVE COWORKING - INVOICE EXCEL TEMPLATE"],
        [""],
        ["HOW TO USE:"],
        ["1. Fill one row per client/invoice"],
        ["2. Only fill the service columns that apply - empty columns are ignored"],
        ["3. GST (18%) is auto-calculated for applicable services"],
        [""],
        ["COLUMN GUIDE:"],
        [""],
        ["CLIENT DETAILS (Required):"],
        ["Client Company Name", "Company name as it should appear on invoice"],
        ["Client Address", "Full address including city, state, PIN"],
        ["Client GSTIN", "GST Identification Number (15 characters)"],
        [""],
        ["DATES (Required):"],
        ["Invoice Date", "Format: YYYY-MM-DD (e.g., 2026-01-01)"],
        ["Due Date", "Format: YYYY-MM-DD - Payment due date"],
        [""],
        ["SERVICES (Fill only what applies - leave others empty):"],
        [""],
        ["Monthly Plan Fee", "GST APPLICABLE - Full monthly desk/cabin rate"],
        ["Prorate Days", "Optional - Days to charge (for partial month)"],
        ["Prorate Total Days", "Optional - Total days in period (default: 30)"],
        ["", "Example: Fee=15000, Days=15, Total=30 → Charges ₹7,500 + GST"],
        [""],
        ["Day Pass Rate", "GST APPLICABLE - Rate per day pass"],
        ["Day Pass Qty", "Number of day passes (default: 1)"],
        [""],
        ["Security Deposit", "NO GST - Refundable security deposit amount"],
        [""],
        ["Setup Charges", "GST APPLICABLE - One-time setup/onboarding fee"],
        [""],
        ["Meeting Room Rate", "GST APPLICABLE - Rate per hour"],
        ["Meeting Room Hours", "Number of hours (default: 1)"],
        [""],
        ["Additional Charges", "GST APPLICABLE - Any other charges"],
        ["Additional Charges Description", "Description for additional charges"],
        [""],
        ["Notes", "Optional - Notes to appear on invoice"],
        [""],
        ["TIPS:"],
        ["• Leave service columns empty if not applicable"],
        ["• Only filled services appear on the invoice"],
        ["• New clients are auto-created in the system"],
        ["• Multiple invoices = multiple rows"],
    ]
    
    for row_num, row_data in enumerate(instructions, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_instructions.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:
                cell.font = Font(bold=True, size=14, color="2E375B")
            elif row_num in [3, 10, 15, 18, 20, 25, 28, 30, 33, 36, 39, 42]:
                cell.font = Font(bold=True)
            elif "GST APPLICABLE" in str(value):
                cell.font = Font(color="008000")
            elif "NO GST" in str(value):
                cell.font = Font(color="FF6600")
    
    ws_instructions.column_dimensions['A'].width = 30
    ws_instructions.column_dimensions['B'].width = 55
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=thryve_invoice_template.xlsx"}
    )

# Process Excel and generate invoices
@api_router.post("/excel/upload")
async def upload_excel_and_generate_invoices(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx or .xls)")
    
    try:
        contents = await file.read()
        wb = load_workbook(filename=BytesIO(contents))
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        created_invoices = []
        errors = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:  # Skip empty rows
                continue
            
            try:
                row_data = dict(zip(headers, row))
                
                # Extract client info
                client_data = {
                    "company_name": str(row_data.get("Client Company Name", "") or "").strip(),
                    "address": str(row_data.get("Client Address", "") or "").strip(),
                    "gstin": str(row_data.get("Client GSTIN", "") or "").strip().upper()
                }
                
                if not client_data["company_name"] or not client_data["gstin"]:
                    errors.append({"row": row_num, "error": "Missing client name or GSTIN"})
                    continue
                
                # Check if client exists, if not create
                client = await db.clients.find_one(
                    {"gstin": client_data["gstin"]}, 
                    {"_id": 0}
                )
                
                if not client:
                    client_obj = Client(**client_data)
                    client_doc = client_obj.model_dump()
                    await db.clients.insert_one(client_doc)
                    client = client_doc
                
                # Build line items from filled columns
                line_items = []
                
                # Monthly Plan Fee
                monthly_fee = row_data.get("Monthly Plan Fee")
                if monthly_fee and str(monthly_fee).strip():
                    monthly_fee = float(monthly_fee)
                    prorate_days = row_data.get("Prorate Days")
                    prorate_total = row_data.get("Prorate Total Days") or 30
                    is_prorated = prorate_days is not None and str(prorate_days).strip() != ""
                    
                    line_items.append(LineItem(
                        description="Monthly Plan Fee" + (f" (Prorated: {prorate_days} of {prorate_total} days)" if is_prorated else ""),
                        service_type="monthly_rental",
                        quantity=1,
                        rate=monthly_fee,
                        is_taxable=True,
                        hsn_sac="997212",
                        unit="Month",
                        is_prorated=is_prorated,
                        prorate_days=int(prorate_days) if is_prorated else None,
                        prorate_total_days=int(prorate_total) if prorate_total else 30
                    ))
                
                # Day Pass
                day_pass_rate = row_data.get("Day Pass Rate")
                if day_pass_rate and str(day_pass_rate).strip():
                    day_pass_qty = row_data.get("Day Pass Qty") or 1
                    line_items.append(LineItem(
                        description="Day Pass",
                        service_type="day_pass",
                        quantity=float(day_pass_qty),
                        rate=float(day_pass_rate),
                        is_taxable=True,
                        hsn_sac="997212",
                        unit="Day"
                    ))
                
                # Security Deposit (NO GST)
                security_deposit = row_data.get("Security Deposit")
                if security_deposit and str(security_deposit).strip():
                    line_items.append(LineItem(
                        description="Refundable Security Deposit",
                        service_type="security_deposit",
                        quantity=1,
                        rate=float(security_deposit),
                        is_taxable=False,
                        hsn_sac="",
                        unit="Units"
                    ))
                
                # Setup Charges
                setup_charges = row_data.get("Setup Charges")
                if setup_charges and str(setup_charges).strip():
                    line_items.append(LineItem(
                        description="Setup Charges",
                        service_type="setup_charges",
                        quantity=1,
                        rate=float(setup_charges),
                        is_taxable=True,
                        hsn_sac="997212",
                        unit="Units"
                    ))
                
                # Meeting Room
                meeting_rate = row_data.get("Meeting Room Rate")
                if meeting_rate and str(meeting_rate).strip():
                    meeting_hours = row_data.get("Meeting Room Hours") or 1
                    line_items.append(LineItem(
                        description="Meeting Room Charges",
                        service_type="meeting_room",
                        quantity=float(meeting_hours),
                        rate=float(meeting_rate),
                        is_taxable=True,
                        hsn_sac="997212",
                        unit="Hour"
                    ))
                
                # Additional Charges
                additional = row_data.get("Additional Charges")
                if additional and str(additional).strip():
                    additional_desc = row_data.get("Additional Charges Description") or "Additional Charges"
                    line_items.append(LineItem(
                        description=str(additional_desc),
                        service_type="monthly_rental",  # Use monthly_rental type for GST
                        quantity=1,
                        rate=float(additional),
                        is_taxable=True,
                        hsn_sac="997212",
                        unit="Units"
                    ))
                
                if not line_items:
                    errors.append({"row": row_num, "client": client_data["company_name"], "error": "No billable items found"})
                    continue
                
                # Calculate line items
                calculated_items = [calculate_line_item(item) for item in line_items]
                
                # Calculate totals
                subtotal = sum(item["amount"] for item in calculated_items)
                total_cgst = sum(item["cgst"] for item in calculated_items)
                total_sgst = sum(item["sgst"] for item in calculated_items)
                total_tax = total_cgst + total_sgst
                grand_total = subtotal + total_tax
                
                # Get dates
                invoice_date = row_data.get("Invoice Date (YYYY-MM-DD)")
                due_date = row_data.get("Due Date (YYYY-MM-DD)")
                
                if invoice_date:
                    invoice_date = str(invoice_date).strip()[:10]
                else:
                    invoice_date = datetime.now().strftime('%Y-%m-%d')
                
                if due_date:
                    due_date = str(due_date).strip()[:10]
                
                # Generate invoice number with company name
                company_name = client_data.get("company_name", "")
                invoice_number = await generate_invoice_number(company_name)
                
                # Create invoice
                invoice_obj = Invoice(
                    invoice_number=invoice_number,
                    invoice_date=invoice_date,
                    due_date=due_date if due_date else None,
                    client=client,
                    line_items=calculated_items,
                    subtotal=round(subtotal, 2),
                    total_cgst=round(total_cgst, 2),
                    total_sgst=round(total_sgst, 2),
                    total_tax=round(total_tax, 2),
                    grand_total=round(grand_total, 2),
                    notes=str(row_data.get("Notes", "") or ""),
                    status="pending"
                )
                
                doc = invoice_obj.model_dump()
                doc = {k: v for k, v in doc.items() if v is not None}
                await db.invoices.insert_one(doc)
                
                created_invoices.append({
                    "invoice_number": invoice_number,
                    "client": client_data["company_name"],
                    "items_count": len(calculated_items),
                    "grand_total": round(grand_total, 2)
                })
                
            except Exception as e:
                errors.append({
                    "row": row_num,
                    "client": row_data.get("Client Company Name", "Unknown"),
                    "error": str(e)
                })
        
        return {
            "success": True,
            "created_count": len(created_invoices),
            "created_invoices": created_invoices,
            "errors": errors
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing Excel file: {str(e)}")

# Include the router in the main app
app.include_router(api_router)

# Include management routes
from routes.management import router as management_router, init_router as init_management_router
init_management_router(db, get_current_user, check_permission)
app.include_router(management_router)

# Include member portal routes
from routes.member_portal import router as member_portal_router, init_router as init_member_portal_router
init_member_portal_router(db)
app.include_router(member_portal_router)

# Include public holidays routes
from routes.public_holidays import router as holidays_router, init_router as init_holidays_router, seed_default_holidays
init_holidays_router(db, get_current_user, check_permission)
app.include_router(holidays_router)

# Include auto invoice routes
from routes.auto_invoice import router as auto_invoice_router, init_router as init_auto_invoice_router
init_auto_invoice_router(db, get_current_user, check_permission, COMPANY_DETAILS)
app.include_router(auto_invoice_router)

# Include company routes
from routes.companies import router as companies_router, init_router as init_companies_router
init_companies_router(db, get_current_user, check_permission)
app.include_router(companies_router)

# Include import routes (for one-time bulk data import)
from routes.import_data import router as import_router, init_router as init_import_router
init_import_router(db, get_current_user, check_permission)
app.include_router(import_router)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
